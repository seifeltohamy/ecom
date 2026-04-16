"""
chainz_export.py — Chainz Solutions export automation

1. Login to partner.chainzsolutions.com
2. Navigate to Orders page
3. Click EXPORT button (triggers email with Excel attachment)
4. Poll Gmail IMAP for the Chainz export email
5. Extract .xlsx attachment and save to temp file
"""

import email as email_lib
import imaplib
import os
import logging
import tempfile
import time
from datetime import datetime

import openpyxl
from playwright.sync_api import sync_playwright

log = logging.getLogger(__name__)


# ── Step 1: Trigger export from Chainz portal ────────────────────────────────

def trigger_chainz_export(email: str, password: str):
    """Login to partner.chainzsolutions.com and click EXPORT on the Orders page."""
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True, args=["--no-sandbox"])
        page = browser.new_page()
        try:
            page.goto("https://partner.chainzsolutions.com", wait_until="networkidle", timeout=30000)
            time.sleep(2)

            # Try common login selectors
            for sel in ['input[type="email"]', 'input[name="email"]', 'input[placeholder*="mail" i]', 'input[type="text"]']:
                if page.locator(sel).count() > 0:
                    page.fill(sel, email)
                    log.info("Email filled using selector: %s", sel)
                    break
            else:
                raise RuntimeError(f"Could not find email input on {page.url}")

            for sel in ['input[type="password"]', 'input[name="password"]']:
                if page.locator(sel).count() > 0:
                    page.fill(sel, password)
                    break
            else:
                raise RuntimeError("Could not find password input")

            # Submit login
            for sel in ['button[type="submit"]', 'button:has-text("Login")', 'button:has-text("Sign")', 'input[type="submit"]']:
                if page.locator(sel).count() > 0:
                    page.click(sel)
                    break

            page.wait_for_load_state("networkidle", timeout=15000)
            time.sleep(2)

            # Navigate to Orders page
            page.goto("https://partner.chainzsolutions.com/orders", wait_until="networkidle", timeout=30000)
            time.sleep(2)

            # Click EXPORT button
            export_btn = page.locator('button:has-text("EXPORT"), a:has-text("EXPORT"), button:has-text("Export")')
            if export_btn.count() > 0:
                export_btn.first.click()
                log.info("EXPORT button clicked")
                time.sleep(3)
            else:
                raise RuntimeError("Could not find EXPORT button on orders page")

        finally:
            browser.close()


# ── Step 2: Fetch export email from Gmail IMAP ──────────────────────────────

def fetch_chainz_email(gmail_user: str, gmail_pass: str, timeout: int = 300) -> str:
    """Poll Gmail IMAP for Chainz export email with .xlsx attachment.
    Returns path to saved attachment."""
    deadline = time.time() + timeout
    since_str = datetime.now().strftime("%d-%b-%Y")
    while time.time() < deadline:
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(gmail_user, gmail_pass)
            mail.select('"[Gmail]/All Mail"')

            # Search for recent emails with "Orders Excel Report" in subject
            # (covers "Orders Excel Report - Zen", "Orders Excel Report - BrandName", etc.)
            _, data = mail.search(None, f'SUBJECT "Orders Excel Report" SINCE "{since_str}"')

            ids = data[0].split()
            if ids:
                # Check from most recent first
                for eid in reversed(ids):
                    _, msg_data = mail.fetch(eid, "(RFC822)")
                    msg = email_lib.message_from_bytes(msg_data[0][1])

                    # Look for .xlsx attachment
                    for part in msg.walk():
                        filename = part.get_filename()
                        if filename and filename.endswith(".xlsx"):
                            tmp = tempfile.mkdtemp()
                            path = os.path.join(tmp, filename)
                            with open(path, "wb") as f:
                                f.write(part.get_payload(decode=True))
                            mail.logout()
                            log.info("Chainz export attachment saved: %s", path)
                            return path

            mail.logout()
        except Exception as e:
            log.warning("IMAP poll error: %s", e)

        time.sleep(15)

    raise RuntimeError("Timed out waiting for Chainz export email")


# ── Step 3: Sort Chainz Excel by Ordered At ──────────────────────────────────

def sort_chainz_excel(path: str) -> tuple[str, str, str]:
    """Sort Chainz Excel by 'Ordered At' column. Returns (output_path, date_from, date_to)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    try:
        date_idx = headers.index("Ordered At")
    except ValueError:
        raise RuntimeError("No 'Ordered At' column found in Chainz Excel")

    data_rows = list(ws.iter_rows(min_row=2, values_only=True))

    def parse_date(row):
        raw = row[date_idx]
        if raw is None:
            return datetime.min
        if isinstance(raw, datetime):
            return raw
        try:
            return datetime.strptime(str(raw).strip(), "%Y-%m-%d %H:%M")
        except ValueError:
            try:
                return datetime.strptime(str(raw).strip(), "%Y-%m-%d")
            except ValueError:
                return datetime.min

    data_rows.sort(key=parse_date)

    dates = [parse_date(r) for r in data_rows if parse_date(r) != datetime.min]
    date_from = min(dates).strftime("%Y-%m-%d") if dates else ""
    date_to = max(dates).strftime("%Y-%m-%d") if dates else ""

    # Write sorted
    out_path = path.replace(".xlsx", "_sorted.xlsx")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(headers)
    for row in data_rows:
        out_ws.append(list(row))
    out_wb.save(out_path)

    return out_path, date_from, date_to
