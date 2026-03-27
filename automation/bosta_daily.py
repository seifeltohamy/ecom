#!/usr/bin/env python3
"""
bosta_daily.py — Daily Bosta export automation

For each brand configured in EcomHQ Settings (bosta_email + bosta_password + bosta_email_password):
  1. Login to business.bosta.co and click Export on the Successful orders tab
     → Bosta emails the Excel file to the bosta_email inbox
  2. Poll Gmail IMAP for the export email → extract download link → download file
  3. Sort ALL rows by "Delivered at" (ascending)
  4. Filter to current month only
  5. Upload sorted+filtered Excel to EcomHQ portal

Config: automation/.env.automation (copy from .env.automation.example)
Logs:   /tmp/bosta_daily.log  (or LOG_FILE in config)

Run manually: .venv/bin/python automation/bosta_daily.py
"""

import email as email_lib
import imaplib
import os
import logging
import re
import tempfile
import time
from datetime import date, datetime

import httpx
import openpyxl
from dotenv import dotenv_values
from playwright.sync_api import sync_playwright

# ── Config ────────────────────────────────────────────────────────────────────

CFG_PATH   = os.path.join(os.path.dirname(__file__), ".env.automation")
cfg        = dotenv_values(CFG_PATH)

ECOMHQ_URL  = cfg.get("ECOMHQ_URL",            "https://ecom-production-a643.up.railway.app")
ADMIN_EMAIL = cfg.get("ECOMHQ_ADMIN_EMAIL",     "")
ADMIN_PASS  = cfg.get("ECOMHQ_ADMIN_PASSWORD",  "")
LOG_FILE    = cfg.get("LOG_FILE",               "/tmp/bosta_daily.log")

log = logging.getLogger(__name__)

# ── Step 1: Fetch brand configs ───────────────────────────────────────────────

def get_brand_configs():
    """Login as admin, return list of brand configs + admin token."""
    r = httpx.post(f"{ECOMHQ_URL}/auth/login",
                   data={"username": ADMIN_EMAIL, "password": ADMIN_PASS},
                   timeout=15)
    r.raise_for_status()
    token = r.json()["access_token"]

    r = httpx.get(f"{ECOMHQ_URL}/admin/brand-settings",
                  headers={"Authorization": f"Bearer {token}"},
                  timeout=15)
    r.raise_for_status()
    return r.json(), token  # [{brand_id, brand_name, bosta_email, bosta_password, bosta_email_password}], token

# ── Step 2a: Trigger export via Playwright ────────────────────────────────────

def trigger_bosta_export(email: str, password: str) -> None:
    """Login to business.bosta.co, click Successful tab, click Export.
    Bosta sends the Excel file to the email inbox — no direct download here."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        ctx = browser.new_context()
        page = ctx.new_page()
        page.set_default_timeout(60000)  # 60s for all actions

        log.info("  Navigating to Bosta signin…")
        page.goto("https://business.bosta.co/signin", timeout=60000)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(2000)  # extra settle time after networkidle

        # Try multiple selectors — Bosta occasionally changes their login page
        EMAIL_SELECTORS = [
            'input[type="email"]',
            'input[name="email"]',
            'input[placeholder*="mail" i]',
            'input[placeholder*="الإيميل"]',
            'input[placeholder*="البريد"]',
            'input[type="text"]',  # last resort: first text input
        ]
        PASSWORD_SELECTORS = [
            'input[type="password"]',
            'input[name="password"]',
        ]

        email_field = None
        for sel in EMAIL_SELECTORS:
            try:
                page.wait_for_selector(sel, state="visible", timeout=5000)
                email_field = sel
                break
            except Exception:
                continue

        if not email_field:
            # Dump page HTML to log for debugging
            log.error(f"  Could not find email input. Page URL: {page.url}")
            log.error(f"  Page title: {page.title()}")
            raise RuntimeError("Email input not found on Bosta signin page — selectors need updating")

        log.info(f"  Using email selector: {email_field}")
        page.fill(email_field, email)
        page.fill(PASSWORD_SELECTORS[0], password)
        page.click('button[type="submit"]')
        page.wait_for_url("**/overview**", timeout=60000)
        page.wait_for_load_state("networkidle")
        log.info(f"  Logged in — current URL: {page.url}")

        log.info("  Navigating to orders page…")
        page.goto("https://business.bosta.co/orders", timeout=60000)
        page.wait_for_load_state("networkidle")
        log.info(f"  Orders page loaded — current URL: {page.url}")

        log.info("  Clicking Successful tab (تم بنجاح)…")
        page.wait_for_selector('text=تم بنجاح', state='visible', timeout=30000)
        page.click('text=تم بنجاح', timeout=30000)
        page.wait_for_load_state("networkidle")

        log.info("  Clicking Export (تحميل)…")
        page.click('button:has-text("تحميل")')
        # Bosta sends file to email — no browser download to capture
        page.wait_for_timeout(3000)

        ctx.close()
        browser.close()
        log.info("  Export triggered — Bosta will email the file.")

# ── Step 2b: Fetch export link from Gmail IMAP ────────────────────────────────

def fetch_export_from_email(gmail_user: str, gmail_app_password: str,
                             timeout: int = 300) -> str:
    """Poll Gmail IMAP for Bosta export email, return download link."""
    from datetime import timezone
    import email.utils

    triggered_at = time.time()

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(gmail_user, gmail_app_password)

    deadline = triggered_at + timeout
    while time.time() < deadline:
        # Search All Mail (catches Inbox, Promotions, Spam, etc.)
        mail.select('"[Gmail]/All Mail"')

        # Broad search — just FROM bosta.co, no subject filter
        _, msgs = mail.search(None, 'FROM "bosta.co"')
        ids = msgs[0].split()
        log.info(f"  IMAP: {len(ids)} email(s) from bosta.co found")

        if ids:
            # Check most recent 10 emails first
            for msg_id in reversed(ids[-10:]):
                _, data = mail.fetch(msg_id, "(RFC822)")
                msg = email_lib.message_from_bytes(data[0][1])

                subject  = msg.get("Subject", "")
                sender   = msg.get("From", "")
                date_str = msg.get("Date", "")
                try:
                    msg_time = email.utils.parsedate_to_datetime(date_str).timestamp()
                except Exception:
                    msg_time = 0

                log.info(f"  Email: from={sender} | subject={subject} | age={int(time.time()-msg_time)}s")

                # Only use emails that arrived after we triggered the export
                if msg_time < triggered_at - 60:
                    continue

                # Extract download link from body
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() in ("text/html", "text/plain"):
                            body += part.get_payload(decode=True).decode(errors="ignore")
                else:
                    body = msg.get_payload(decode=True).decode(errors="ignore")

                links = re.findall(r'https?://[^\s"<>\']+', body)
                for link in links:
                    if "download" in link.lower() or "export" in link.lower() or "storage" in link.lower():
                        log.info(f"  Found export link: {link[:80]}…")
                        mail.logout()
                        return link

        log.info("  Export email not yet received, waiting 15s…")
        time.sleep(15)

    mail.logout()
    raise TimeoutError("Bosta export email not received within timeout")


def download_from_link(link: str, download_dir: str) -> str:
    """Download file from URL, return local path."""
    r = httpx.get(link, follow_redirects=True, timeout=60)
    r.raise_for_status()
    filename = link.split("/")[-1].split("?")[0] or "bosta_export.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    path = os.path.join(download_dir, filename)
    with open(path, "wb") as f:
        f.write(r.content)
    log.info(f"  File saved → {path}")
    return path

# ── Step 3: Sort then Filter ──────────────────────────────────────────────────

def parse_delivered_at(val) -> date | None:
    if not val:
        return None
    s = str(val).strip()
    for fmt in ("%m-%d-%Y, %H:%M:%S", "%m-%d-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def sort_then_filter(path: str) -> tuple[str, str, str]:
    """Sort ALL rows by 'Delivered at' ascending, then filter to current month.
    Returns (output_path, date_from_iso, date_to_iso)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_idx = next((i for i, h in enumerate(header) if h == "Delivered at"), None)
    if col_idx is None:
        raise ValueError(f"'Delivered at' column not found. Headers: {header}")

    today = date.today()
    first = today.replace(day=1)

    data = list(rows[1:])
    data.sort(key=lambda r: parse_delivered_at(r[col_idx]) or date.min)

    filtered = [
        r for r in data
        if (d := parse_delivered_at(r[col_idx])) and first <= d <= today
    ]

    log.info(f"  Rows before filter: {len(data)} → after filter: {len(filtered)}")

    ws.delete_rows(2, ws.max_row)
    for row in filtered:
        ws.append(list(row))

    out = path.replace(".xlsx", "_sorted_filtered.xlsx")
    wb.save(out)
    return out, first.isoformat(), today.isoformat()

def sort_only(path: str) -> tuple[str, str, str]:
    """Sort all rows by 'Delivered at' ascending (no month filter).
    Returns (output_path, min_date_iso, max_date_iso)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    col_idx = next((i for i, h in enumerate(header) if h == "Delivered at"), None)
    if col_idx is None:
        raise ValueError(f"'Delivered at' column not found. Headers: {header}")

    data = sorted(rows[1:], key=lambda r: parse_delivered_at(r[col_idx]) or date.min)
    dates = [parse_delivered_at(r[col_idx]) for r in data if parse_delivered_at(r[col_idx])]
    min_date = dates[0].isoformat()  if dates else date.today().replace(day=1).isoformat()
    max_date = dates[-1].isoformat() if dates else date.today().isoformat()

    ws.delete_rows(2, ws.max_row)
    for row in data:
        ws.append(list(row))

    out = path.replace(".xlsx", "_sorted.xlsx")
    wb.save(out)
    log.info(f"  Sorted {len(data)} rows: {min_date} → {max_date}")
    return out, min_date, max_date


# ── Step 4: Upload to EcomHQ ──────────────────────────────────────────────────

def set_meta_ads_spent(brand_token: str, report_id: int,
                       date_from: str, date_to: str) -> None:
    """Fetch Meta spend for the report's date range and save it on the report."""
    try:
        r = httpx.get(
            f"{ECOMHQ_URL}/meta/summary",
            headers={"Authorization": f"Bearer {brand_token}"},
            params={"date_from": date_from, "date_to": date_to},
            timeout=15,
        )
        if r.status_code != 200:
            log.info(f"  Meta summary: HTTP {r.status_code} — skipping ads_spent")
            return
        data = r.json()
        spend = data.get("spend", 0)
        if not spend:
            log.info("  Meta summary: no spend data for this period — skipping ads_spent")
            return
        httpx.put(
            f"{ECOMHQ_URL}/reports/{report_id}/pl",
            headers={"Authorization": f"Bearer {brand_token}"},
            json={"ads_spent": spend, "items": []},
            timeout=15,
        )
        log.info(f"  Meta ads_spent auto-set: {spend} EGP")
    except Exception as e:
        log.warning(f"  Meta ads_spent auto-set failed (non-fatal): {e}")


def upload_to_ecomhq(admin_token: str, brand_id: int,
                     file_path: str, date_from: str, date_to: str) -> tuple[dict, str]:
    """Returns (report_dict, brand_token)."""
    r = httpx.post(f"{ECOMHQ_URL}/auth/select-brand",
                   json={"brand_id": brand_id},
                   headers={"Authorization": f"Bearer {admin_token}"},
                   timeout=15)
    r.raise_for_status()
    brand_token = r.json()["access_token"]

    with open(file_path, "rb") as f:
        r = httpx.post(
            f"{ECOMHQ_URL}/upload",
            headers={"Authorization": f"Bearer {brand_token}"},
            data={"date_from": date_from, "date_to": date_to},
            files={"file": (
                os.path.basename(file_path), f,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            timeout=120,
        )
    r.raise_for_status()
    return r.json(), brand_token

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    log.info("=== bosta_daily.py started ===")

    if not ADMIN_EMAIL or not ADMIN_PASS:
        log.error("ECOMHQ_ADMIN_EMAIL / ECOMHQ_ADMIN_PASSWORD not set in .env.automation")
        return

    try:
        brands, admin_token = get_brand_configs()
    except Exception as e:
        log.exception(f"Failed to fetch brand configs: {e}")
        return

    if not brands:
        log.info("No brands with Bosta credentials configured — nothing to do.")
        return

    log.info(f"Processing {len(brands)} brand(s)…")

    for brand in brands:
        name           = brand["brand_name"]
        brand_id       = brand["brand_id"]
        email          = brand["bosta_email"]
        password       = brand["bosta_password"]
        email_password = brand.get("bosta_email_password", "")

        if not email_password:
            log.warning(f"  Brand '{name}': no bosta_email_password set — skipping.")
            continue

        log.info(f"--- Brand: {name} (id={brand_id}) ---")
        try:
            with tempfile.TemporaryDirectory() as tmp:
                trigger_bosta_export(email, password)
                link = fetch_export_from_email(email, email_password)
                path = download_from_link(link, tmp)
                sorted_path, date_from, date_to = sort_then_filter(path)
                result, brand_token = upload_to_ecomhq(admin_token, brand_id, sorted_path, date_from, date_to)
                log.info(
                    f"  Uploaded OK — report_id={result.get('report_id')}, "
                    f"orders={result.get('order_count')}, "
                    f"revenue={result.get('grand_revenue')}"
                )
                # Auto-fill ads_spent from Meta API (non-fatal if brand not connected)
                report_id = result.get("report_id") or result.get("id")
                if report_id:
                    set_meta_ads_spent(brand_token, report_id, date_from, date_to)
        except Exception as e:
            log.exception(f"  Brand '{name}' failed: {e}")

    log.info("=== bosta_daily.py done ===")


if __name__ == "__main__":
    main()
