"""
excel_helpers.py — Shared Excel parsing, report building, and export session helpers.

Used by bosta.py, reports.py, and automation.py routers.
"""

import os
import re
import io
import json
from collections import defaultdict
from datetime import datetime as _dt

import openpyxl
from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.deps import get_db
from app import models


# ── Export session persistence (disk-based for multi-worker) ─────────────────

def export_meta_path(fid: str) -> str:
    return f"/tmp/ecomhq_export_{fid}.json"

def save_export(fid: str, info: dict) -> None:
    with open(export_meta_path(fid), "w") as f:
        json.dump(info, f)

def load_export(fid: str) -> dict | None:
    p = export_meta_path(fid)
    if not os.path.exists(p):
        return None
    try:
        with open(p) as f:
            return json.load(f)
    except Exception:
        return None

def delete_export(fid: str) -> None:
    try:
        os.remove(export_meta_path(fid))
    except FileNotFoundError:
        pass


# ── Product helpers ──────────────────────────────────────────────────────────

def get_products_map(db: Session, brand_id: int) -> dict:
    """Return {sku: name} for backward compat."""
    products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
    return {p.sku: p.name for p in products}


def get_products_full(db: Session, brand_id: int) -> dict:
    """Return {sku: {name, price}} for price-aware report building."""
    products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
    return {p.sku: {"name": p.name, "price": p.price} for p in products}


# ── Bosta Excel parsing ─────────────────────────────────────────────────────

def parse_description_text(text: str) -> list[tuple]:
    """Extract (sku, quantity, price) tuples from a Bosta Description cell."""
    pattern = r"BostaSKU:(BO-\d+)\s*-\s*quantity:(\d+)\s*-\s*itemPrice:([\d.]+)"
    return re.findall(pattern, text)


def aggregate_excel(workbook, date_from: str = None, date_to: str = None) -> tuple[dict, int]:
    """Parse Bosta Excel into sku_data dict. Returns (sku_data, order_count)."""
    from datetime import datetime, date

    ws = workbook.active
    headers = [cell.value for cell in ws[1]]

    try:
        desc_idx = headers.index("Description")
    except ValueError:
        raise HTTPException(status_code=400, detail="No 'Description' column found in Excel.")

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    if "delivered at" in headers_lower:
        deliv_idx = headers_lower.index("delivered at")
    else:
        deliv_idx = next(
            (i for i, h in enumerate(headers_lower) if h == "delivered at"),
            None
        )

    dt_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").date() if date_to   else None

    sku_data = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "total": 0.0}))
    order_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if deliv_idx is not None and (dt_from or dt_to):
            raw = row[deliv_idx]
            if raw is None:
                continue
            if isinstance(raw, (datetime, date)):
                row_date = raw.date() if isinstance(raw, datetime) else raw
            else:
                raw_str = str(raw).strip()
                parsed = None
                for fmt in ("%m-%d-%Y, %H:%M:%S", "%m-%d-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                row_date = parsed
            if dt_from and row_date < dt_from:
                continue
            if dt_to   and row_date > dt_to:
                continue

        cell = row[desc_idx]
        if not cell:
            continue
        matches = parse_description_text(str(cell))
        if matches:
            order_count += 1
            for sku, qty_str, price_str in matches:
                qty = int(qty_str)
                price = float(price_str)
                sku_data[sku][price]["quantity"] += qty
                sku_data[sku][price]["total"] += qty * price

    return sku_data, order_count


# ── Chainz Excel parsing ────────────────────────────────────────────────────

def parse_chainz_items(items_str: str) -> list[tuple[str, int]]:
    """Parse Chainz Items column: 'Zen-1011: 1 pcs, Zen-1010: 3 pcs' → [(sku, qty), ...]"""
    if not items_str:
        return []
    return [(m.group(1), int(m.group(2))) for m in re.finditer(r"(\S+?):\s*(\d+)\s*pcs", str(items_str))]


def aggregate_chainz_excel(workbook, date_from: str = None, date_to: str = None,
                           price_lookup: dict = None) -> tuple[dict, int]:
    """Parse Chainz Excel into same sku_data shape as Bosta."""
    from datetime import datetime, date

    ws = workbook.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    ordered_at_idx = col("Ordered At")
    status_idx = col("Status")
    items_idx = col("Items")
    order_num_idx = col("Shopify Number") or col("Shopify Order #")
    total_items_idx = col("Total Items")
    cod_idx = col("COD")

    if items_idx is None:
        raise HTTPException(status_code=400, detail="No 'Items' column found in Chainz Excel.")

    dt_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    dt_to = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else None

    sku_data = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "total": 0.0}))
    order_count = 0
    price_lookup = price_lookup or {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if status_idx is not None:
            status = str(row[status_idx] or "").strip()
            if status != "Delivered":
                continue

        if ordered_at_idx is not None and (dt_from or dt_to):
            raw = row[ordered_at_idx]
            if raw is None:
                continue
            if isinstance(raw, (datetime, date)):
                row_date = raw.date() if isinstance(raw, datetime) else raw
            else:
                raw_str = str(raw).strip()
                parsed = None
                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                row_date = parsed
            if dt_from and row_date < dt_from:
                continue
            if dt_to and row_date > dt_to:
                continue

        items = parse_chainz_items(row[items_idx] if items_idx is not None else "")
        if not items:
            continue

        order_count += 1
        order_name = str(row[order_num_idx] or "").strip() if order_num_idx is not None else ""
        order_prices = price_lookup.get(order_name, {})

        cod = float(row[cod_idx] or 0) if cod_idx is not None else 0
        total_items_val = int(row[total_items_idx] or 0) if total_items_idx is not None else 0
        fallback_price = (cod / total_items_val) if total_items_val > 0 else 0

        for sku, qty in items:
            price = order_prices.get(sku, fallback_price)
            sku_data[sku][price]["quantity"] += qty
            sku_data[sku][price]["total"] += qty * price

    return sku_data, order_count


# ── Report builder ───────────────────────────────────────────────────────────

def build_report(sku_data: dict, products: dict, order_count: int = 0, product_prices: dict = None) -> dict:
    """Build report from sku_data. products = {sku: name}, product_prices = {sku: float|None}."""
    rows = []
    grand_qty = 0
    grand_rev = 0.0
    product_prices = product_prices or {}

    product_order = {sku: i for i, sku in enumerate(products.keys())}

    def sort_key(sku):
        return (product_order.get(sku, 9999), -sum(v["total"] for v in sku_data[sku].values()))

    for sku in sorted(sku_data.keys(), key=sort_key):
        name = products.get(sku, "Unknown Product")
        saved_price = product_prices.get(sku)

        if saved_price is not None:
            total_qty = sum(d["quantity"] for d in sku_data[sku].values())
            total_rev = round(saved_price * total_qty, 2)
            price_breakdown = [{"price": saved_price, "quantity": total_qty, "total": total_rev}]
            sku_qty = total_qty
            sku_rev = total_rev
        else:
            price_breakdown = []
            sku_qty = 0
            sku_rev = 0.0
            for price in sorted(sku_data[sku].keys(), reverse=True):
                d = sku_data[sku][price]
                price_breakdown.append({
                    "price": price,
                    "quantity": d["quantity"],
                    "total": round(d["total"], 2),
                })
                sku_qty += d["quantity"]
                sku_rev += d["total"]

        rows.append({
            "sku": sku,
            "name": name,
            "prices": price_breakdown,
            "total_quantity": sku_qty,
            "total_revenue": round(sku_rev, 2),
        })

        grand_qty += sku_qty
        grand_rev += sku_rev

    return {
        "rows": rows,
        "grand_quantity": grand_qty,
        "grand_revenue": round(grand_rev, 2),
        "order_count": order_count,
    }


# ── Process Excel (shared by upload and automation) ──────────────────────────

async def process_excel(contents: bytes, date_from: str | None, date_to: str | None, brand_id: int, provider: str = "bosta") -> dict:
    """Core Excel processing logic shared by /upload and /automation/upload/{file_id}."""
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    if provider == "chainz":
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        order_num_idx = None
        for name in ("Shopify Number", "Shopify Order #"):
            if name in headers:
                order_num_idx = headers.index(name)
                break
        order_names = set()
        if order_num_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = str(row[order_num_idx] or "").strip()
                if val:
                    order_names.add(val)

        price_lookup = {}
        if order_names:
            with get_db() as db:
                settings = {r.key: r.value for r in db.query(models.AppSettings).filter(
                    models.AppSettings.brand_id == brand_id).all()}
            store_url = settings.get("shopify_store_url", "")
            access_token = settings.get("shopify_access_token", "")
            if store_url and access_token:
                from app.shopify_client import get_orders_by_name
                try:
                    price_lookup = get_orders_by_name(store_url, access_token, list(order_names))
                except Exception:
                    pass

        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sku_data, order_count = aggregate_chainz_excel(wb, date_from=date_from, date_to=date_to, price_lookup=price_lookup)
    else:
        sku_data, order_count = aggregate_excel(wb, date_from=date_from, date_to=date_to)

    with get_db() as db:
        products = get_products_map(db, brand_id)

    # Auto-fetch missing or unnamed product names from Shopify and save to DB
    missing_skus = [sku for sku in sku_data if sku not in products or products.get(sku) == "Unknown Product"]
    if missing_skus:
        with get_db() as db:
            settings = {r.key: r.value for r in db.query(models.AppSettings).filter(
                models.AppSettings.brand_id == brand_id).all()}
        store_url = settings.get("shopify_store_url", "")
        access_token = settings.get("shopify_access_token", "")
        if store_url and access_token:
            try:
                from app.shopify_client import get_product_names_by_sku
                shopify_names = get_product_names_by_sku(store_url, access_token)
                with get_db() as db:
                    for sku in missing_skus:
                        name = shopify_names.get(sku)
                        if name:
                            existing = db.query(models.Product).filter(
                                models.Product.sku == sku, models.Product.brand_id == brand_id
                            ).first()
                            if existing:
                                if existing.name == "Unknown Product":
                                    existing.name = name
                            else:
                                db.add(models.Product(sku=sku, name=name, brand_id=brand_id))
                            products[sku] = name
                    db.commit()
            except Exception:
                pass

    # Get saved product prices for price overrides
    with get_db() as db:
        product_prices = {p.sku: p.price for p in db.query(models.Product).filter(
            models.Product.brand_id == brand_id, models.Product.price.isnot(None)
        ).all()}
    report = build_report(sku_data, products, order_count, product_prices=product_prices)
    with get_db() as db:
        saved = models.BostaReport(
            uploaded_at    = _dt.utcnow(),
            date_from      = date_from or None,
            date_to        = date_to   or None,
            order_count    = report["order_count"],
            grand_quantity = report["grand_quantity"],
            grand_revenue  = report["grand_revenue"],
            rows_json      = json.dumps(report["rows"]),
            brand_id       = brand_id,
        )
        db.add(saved)
        db.commit()
        db.refresh(saved)
        report["report_id"] = saved.id
        report["date_from"]  = saved.date_from
        report["date_to"]    = saved.date_to
    return report
