import os
import re
import io
import json
import sys
import uuid
import queue
import shutil
import tempfile
import threading
from collections import defaultdict
from datetime import datetime as _dt
from pathlib import Path

import openpyxl
from pydantic import BaseModel
from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import JSONResponse, StreamingResponse

from app.deps import get_db, get_current_user, get_brand_id, require_writable
from app import models
from sqlalchemy.orm import Session

router = APIRouter()

PROJECT_DIR = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_DIR / "automation"))

# Export sessions stored on disk so all uvicorn workers can share them.
# Each session: /tmp/ecomhq_export_<file_id>.json

def _export_meta_path(fid: str) -> str:
    return f"/tmp/ecomhq_export_{fid}.json"

def _save_export(fid: str, info: dict) -> None:
    with open(_export_meta_path(fid), "w") as f:
        json.dump(info, f)

def _load_export(fid: str) -> dict | None:
    p = _export_meta_path(fid)
    if not os.path.exists(p):
        return None
    try:
        with open(p) as f:
            return json.load(f)
    except Exception:
        return None

def _delete_export(fid: str) -> None:
    try:
        os.remove(_export_meta_path(fid))
    except FileNotFoundError:
        pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_products_map(db: Session, brand_id: int) -> dict:
    """Return {sku: name} for backward compat. Use get_products_full for prices."""
    products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
    return {p.sku: p.name for p in products}


def get_products_full(db: Session, brand_id: int) -> dict:
    """Return {sku: {name, price}} for price-aware report building."""
    products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
    return {p.sku: {"name": p.name, "price": p.price} for p in products}


def parse_description_text(text: str) -> list[tuple]:
    """Extract (sku, quantity, price) tuples from a Description cell."""
    pattern = r"BostaSKU:(BO-\d+)\s*-\s*quantity:(\d+)\s*-\s*itemPrice:([\d.]+)"
    return re.findall(pattern, text)


def aggregate_excel(workbook, date_from: str = None, date_to: str = None) -> dict:
    from datetime import datetime, date

    ws = workbook.active
    headers = [cell.value for cell in ws[1]]

    try:
        desc_idx = headers.index("Description")
    except ValueError:
        raise HTTPException(status_code=400, detail="No 'Description' column found in Excel.")

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    if "delivered at" in headers_lower:
        deliv_idx = headers_lower.index("delivered at")
    else:
        deliv_idx = next(
            (i for i, h in enumerate(headers_lower) if h == "delivered at"),
            None
        )

    dt_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").date() if date_to   else None

    sku_data = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "total": 0.0}))
    order_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if deliv_idx is not None and (dt_from or dt_to):
            raw = row[deliv_idx]
            if raw is None:
                continue
            if isinstance(raw, (datetime, date)):
                row_date = raw.date() if isinstance(raw, datetime) else raw
            else:
                raw_str = str(raw).strip()
                parsed = None
                for fmt in ("%m-%d-%Y, %H:%M:%S", "%m-%d-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                row_date = parsed
            if dt_from and row_date < dt_from:
                continue
            if dt_to   and row_date > dt_to:
                continue

        cell = row[desc_idx]
        if not cell:
            continue
        matches = parse_description_text(str(cell))
        if matches:
            order_count += 1
            for sku, qty_str, price_str in matches:
                qty = int(qty_str)
                price = float(price_str)
                sku_data[sku][price]["quantity"] += qty
                sku_data[sku][price]["total"] += qty * price

    return sku_data, order_count


def parse_chainz_items(items_str: str) -> list[tuple[str, int]]:
    """Parse Chainz Items column: 'Zen-1011: 1 pcs, Zen-1010: 3 pcs' → [(sku, qty), ...]"""
    if not items_str:
        return []
    return [(m.group(1), int(m.group(2))) for m in re.finditer(r"(\S+?):\s*(\d+)\s*pcs", str(items_str))]


def aggregate_chainz_excel(workbook, date_from: str = None, date_to: str = None,
                           price_lookup: dict = None) -> tuple[dict, int]:
    """Parse Chainz Excel into same sku_data shape as Bosta.

    price_lookup: {order_name: {sku: price}} from Shopify API.
    If an order isn't in the lookup, fallback to COD / total_items.
    """
    from datetime import datetime, date

    ws = workbook.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    ordered_at_idx = col("Ordered At")
    status_idx = col("Status")
    items_idx = col("Items")
    order_num_idx = col("Shopify Number") or col("Shopify Order #")
    total_items_idx = col("Total Items")
    cod_idx = col("COD")

    if items_idx is None:
        raise HTTPException(status_code=400, detail="No 'Items' column found in Chainz Excel.")

    dt_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    dt_to = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else None

    sku_data = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "total": 0.0}))
    order_count = 0
    price_lookup = price_lookup or {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Filter by Delivered status
        if status_idx is not None:
            status = str(row[status_idx] or "").strip()
            if status != "Delivered":
                continue

        # Date filtering on "Ordered At"
        if ordered_at_idx is not None and (dt_from or dt_to):
            raw = row[ordered_at_idx]
            if raw is None:
                continue
            if isinstance(raw, (datetime, date)):
                row_date = raw.date() if isinstance(raw, datetime) else raw
            else:
                raw_str = str(raw).strip()
                parsed = None
                for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                row_date = parsed
            if dt_from and row_date < dt_from:
                continue
            if dt_to and row_date > dt_to:
                continue

        items = parse_chainz_items(row[items_idx] if items_idx is not None else "")
        if not items:
            continue

        order_count += 1
        order_name = str(row[order_num_idx] or "").strip() if order_num_idx is not None else ""
        order_prices = price_lookup.get(order_name, {})

        # Fallback: COD / total_items if order not in Shopify lookup
        cod = float(row[cod_idx] or 0) if cod_idx is not None else 0
        total_items_val = int(row[total_items_idx] or 0) if total_items_idx is not None else 0
        fallback_price = (cod / total_items_val) if total_items_val > 0 else 0

        for sku, qty in items:
            price = order_prices.get(sku, fallback_price)
            sku_data[sku][price]["quantity"] += qty
            sku_data[sku][price]["total"] += qty * price

    return sku_data, order_count


def build_report(sku_data: dict, products: dict, order_count: int = 0, product_prices: dict = None) -> dict:
    """Build report from sku_data. products = {sku: name}, product_prices = {sku: float|None}."""
    rows = []
    grand_qty = 0
    grand_rev = 0.0
    product_prices = product_prices or {}

    product_order = {sku: i for i, sku in enumerate(products.keys())}

    def sort_key(sku):
        return (product_order.get(sku, 9999), -sum(v["total"] for v in sku_data[sku].values()))

    for sku in sorted(sku_data.keys(), key=sort_key):
        name = products.get(sku, "Unknown Product")
        saved_price = product_prices.get(sku)

        if saved_price is not None:
            # Override: collapse all price tiers into one using the saved price
            total_qty = sum(d["quantity"] for d in sku_data[sku].values())
            total_rev = round(saved_price * total_qty, 2)
            price_breakdown = [{"price": saved_price, "quantity": total_qty, "total": total_rev}]
            sku_qty = total_qty
            sku_rev = total_rev
        else:
            price_breakdown = []
            sku_qty = 0
            sku_rev = 0.0
            for price in sorted(sku_data[sku].keys(), reverse=True):
                d = sku_data[sku][price]
                price_breakdown.append({
                    "price": price,
                    "quantity": d["quantity"],
                    "total": round(d["total"], 2),
                })
                sku_qty += d["quantity"]
                sku_rev += d["total"]

        rows.append({
            "sku": sku,
            "name": name,
            "prices": price_breakdown,
            "total_quantity": sku_qty,
            "total_revenue": round(sku_rev, 2),
        })

        grand_qty += sku_qty
        grand_rev += sku_rev

    return {
        "rows": rows,
        "grand_quantity": grand_qty,
        "grand_revenue": round(grand_rev, 2),
        "order_count": order_count,
    }


async def _process_excel(contents: bytes, date_from: str | None, date_to: str | None, brand_id: int, provider: str = "bosta") -> dict:
    """Core Excel processing logic shared by /upload and /automation/upload/{file_id}."""
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    if provider == "chainz":
        # Extract Shopify order names from Excel for price lookup
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        order_num_idx = None
        for name in ("Shopify Number", "Shopify Order #"):
            if name in headers:
                order_num_idx = headers.index(name)
                break
        order_names = set()
        if order_num_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = str(row[order_num_idx] or "").strip()
                if val:
                    order_names.add(val)

        # Fetch prices from Shopify
        price_lookup = {}
        if order_names:
            with get_db() as db:
                settings = {r.key: r.value for r in db.query(models.AppSettings).filter(
                    models.AppSettings.brand_id == brand_id).all()}
            store_url = settings.get("shopify_store_url", "")
            access_token = settings.get("shopify_access_token", "")
            if store_url and access_token:
                from app.shopify_client import get_orders_by_name
                try:
                    price_lookup = get_orders_by_name(store_url, access_token, list(order_names))
                except Exception:
                    pass  # fallback to COD / total_items

        # Re-open workbook (iter_rows consumes read-only workbook)
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sku_data, order_count = aggregate_chainz_excel(wb, date_from=date_from, date_to=date_to, price_lookup=price_lookup)
    else:
        sku_data, order_count = aggregate_excel(wb, date_from=date_from, date_to=date_to)
    with get_db() as db:
        products = get_products_map(db, brand_id)

    # Auto-fetch missing product names from Shopify and save to DB
    missing_skus = [sku for sku in sku_data if sku not in products]
    if missing_skus:
        with get_db() as db:
            settings = {r.key: r.value for r in db.query(models.AppSettings).filter(
                models.AppSettings.brand_id == brand_id).all()}
        store_url = settings.get("shopify_store_url", "")
        access_token = settings.get("shopify_access_token", "")
        if store_url and access_token:
            try:
                from app.shopify_client import get_product_names_by_sku
                shopify_names = get_product_names_by_sku(store_url, access_token)
                with get_db() as db:
                    for sku in missing_skus:
                        name = shopify_names.get(sku)
                        if name:
                            existing = db.query(models.Product).filter(
                                models.Product.sku == sku, models.Product.brand_id == brand_id
                            ).first()
                            if not existing:
                                db.add(models.Product(sku=sku, name=name, brand_id=brand_id))
                                products[sku] = name
                    db.commit()
            except Exception:
                pass  # graceful fallback — "Unknown Product" for unfetchable SKUs

    # Get saved product prices for price overrides
    with get_db() as db:
        product_prices = {p.sku: p.price for p in db.query(models.Product).filter(
            models.Product.brand_id == brand_id, models.Product.price.isnot(None)
        ).all()}
    report = build_report(sku_data, products, order_count, product_prices=product_prices)
    with get_db() as db:
        saved = models.BostaReport(
            uploaded_at    = _dt.utcnow(),
            date_from      = date_from or None,
            date_to        = date_to   or None,
            order_count    = report["order_count"],
            grand_quantity = report["grand_quantity"],
            grand_revenue  = report["grand_revenue"],
            rows_json      = json.dumps(report["rows"]),
            brand_id       = brand_id,
        )
        db.add(saved)
        db.commit()
        db.refresh(saved)
        report["report_id"] = saved.id
        report["date_from"]  = saved.date_from
        report["date_to"]    = saved.date_to
    return report


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/debug-upload")
async def debug_upload(file: UploadFile = File(...)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    deliv_idx = headers_lower.index("delivered at") if "delivered at" in headers_lower else None

    samples = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= 10:
            break
        raw = row[deliv_idx] if deliv_idx is not None else "column not found"
        samples.append({"raw": str(raw), "type": type(raw).__name__})

    return {"headers": headers, "delivered_at_index": deliv_idx, "samples": samples}


@router.post("/upload/prepare")
async def upload_prepare(
    file: UploadFile = File(...),
    provider: str = Form("bosta"),
    brand_id: int = Depends(get_brand_id),
    _user: models.User = Depends(require_writable),
):
    """Sort the uploaded file and detect its date range without processing it yet."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls).")
    contents = await file.read()
    tmp = tempfile.mkdtemp()
    path = os.path.join(tmp, "upload.xlsx")
    with open(path, "wb") as f:
        f.write(contents)

    if provider == "chainz":
        # For Chainz, detect date range from "Ordered At" column
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            idx = headers.index("Ordered At") if "Ordered At" in headers else None
            dates = []
            if idx is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    raw = row[idx]
                    if raw is None:
                        continue
                    if isinstance(raw, _dt):
                        dates.append(raw.date())
                    else:
                        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                            try:
                                dates.append(_dt.strptime(str(raw).strip(), fmt).date())
                                break
                            except ValueError:
                                continue
            date_from = min(dates).strftime("%Y-%m-%d") if dates else None
            date_to = max(dates).strftime("%Y-%m-%d") if dates else None
        except Exception as e:
            shutil.rmtree(tmp, ignore_errors=True)
            raise HTTPException(status_code=400, detail=str(e))
    else:
        try:
            import bosta_daily as bd
            path, date_from, date_to = bd.sort_only(path)
        except Exception as e:
            shutil.rmtree(tmp, ignore_errors=True)
            raise HTTPException(status_code=400, detail=str(e))

    fid = str(uuid.uuid4())
    _save_export(fid, {"path": path, "tmp": tmp, "brand_id": brand_id,
                       "date_from": date_from, "date_to": date_to, "provider": provider})
    return {"file_id": fid, "date_from": date_from, "date_to": date_to}


@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    date_from: str = Form(None),
    date_to:   str = Form(None),
    provider:  str = Form("bosta"),
    brand_id: int = Depends(get_brand_id),
    _user: models.User = Depends(require_writable),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls).")
    contents = await file.read()
    report = await _process_excel(contents, date_from, date_to, brand_id, provider=provider)
    return JSONResponse(report)


# ── Reports ───────────────────────────────────────────────────────────────────

@router.get("/reports")
def list_reports(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        reports = db.query(models.BostaReport).filter(
            models.BostaReport.brand_id == brand_id
        ).order_by(models.BostaReport.uploaded_at.desc()).all()
        return [
            {
                "id": r.id,
                "uploaded_at": r.uploaded_at.isoformat(),
                "date_from": r.date_from,
                "date_to": r.date_to,
                "order_count": r.order_count,
                "grand_quantity": r.grand_quantity,
                "grand_revenue": r.grand_revenue,
            }
            for r in reports
        ]


@router.delete("/reports/{report_id}")
def delete_report(report_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(require_writable)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        db.delete(r)
        db.commit()
    return {"ok": True}


@router.get("/reports/{report_id}")
def get_report(report_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        rows = json.loads(r.rows_json)
        # Overlay saved product names so inline-named "Unknown Product" rows
        # reflect the name the user saved even after a refresh / re-upload.
        products_map = get_products_map(db, brand_id)
        for row in rows:
            saved = products_map.get(row.get("sku"))
            if saved:
                row["name"] = saved
        return {
            "report_id": r.id,
            "id": r.id,
            "uploaded_at": r.uploaded_at.isoformat(),
            "date_from": r.date_from,
            "date_to": r.date_to,
            "order_count": r.order_count,
            "grand_quantity": r.grand_quantity,
            "grand_revenue": r.grand_revenue,
            "rows": rows,
        }


@router.get("/reports/{report_id}/pl")
def get_report_pl(report_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        items = db.query(models.BostaReportPl).filter(
            models.BostaReportPl.report_id == report_id
        ).all()
        return {
            "ads_spent": r.ads_spent,
            "items": [
                {
                    "sku": i.sku, "price": i.price,
                    "cost": i.cost, "extra_cost": i.extra_cost,
                    "cost_formula": i.cost_formula, "extra_cost_formula": i.extra_cost_formula,
                }
                for i in items
            ],
        }


class ReportPlItem(BaseModel):
    sku:                str
    price:              float | None = None
    cost:               float | None = None
    extra_cost:         float | None = None
    cost_formula:       str   | None = None
    extra_cost_formula: str   | None = None

class ReportPlPayload(BaseModel):
    ads_spent: float | None = None
    items:     list[ReportPlItem] = []


@router.put("/reports/{report_id}/pl")
def save_report_pl(report_id: int, payload: ReportPlPayload, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(require_writable)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        r.ads_spent = payload.ads_spent
        for item in payload.items:
            row = db.query(models.BostaReportPl).filter(
                models.BostaReportPl.report_id == report_id,
                models.BostaReportPl.sku == item.sku,
            ).first()
            if row:
                row.price              = item.price
                row.cost               = item.cost
                row.extra_cost         = item.extra_cost
                row.cost_formula       = item.cost_formula
                row.extra_cost_formula = item.extra_cost_formula
            else:
                db.add(models.BostaReportPl(
                    report_id=report_id, sku=item.sku,
                    price=item.price, cost=item.cost, extra_cost=item.extra_cost,
                    cost_formula=item.cost_formula, extra_cost_formula=item.extra_cost_formula,
                ))
        db.commit()
    return {"ok": True}


# ── Automation ────────────────────────────────────────────────────────────────

@router.get("/automation/run-export")
def run_export_sse(
    _user:    models.User = Depends(get_current_user),
    brand_id: int = Depends(get_brand_id),
):
    """Stream Bosta export automation progress as SSE. Reads credentials from app_settings."""
    with get_db() as db:
        def _s(key):
            r = db.query(models.AppSettings).filter_by(key=key, brand_id=brand_id).first()
            return r.value if r else None
        bosta_email = _s("bosta_email")
        bosta_pass  = _s("bosta_password")
        email_pass  = _s("bosta_email_password")

    if not all([bosta_email, bosta_pass, email_pass]):
        raise HTTPException(400, "Bosta credentials not configured in Settings")

    q = queue.Queue()

    def _run():
        import bosta_daily as bd
        tmp = tempfile.mkdtemp()
        try:
            q.put("LOG:Logging in to Bosta…")
            bd.trigger_bosta_export(bosta_email, bosta_pass)
            q.put("LOG:Export triggered — waiting for email (up to 5 min)…")
            link = bd.fetch_export_from_email(bosta_email, email_pass)
            q.put("LOG:Download link found — downloading file…")
            path = bd.download_from_link(link, tmp)
            q.put("LOG:File downloaded — sorting rows…")
            out, date_from, date_to = bd.sort_only(path)
            fid = str(uuid.uuid4())
            _save_export(fid, {
                "path": out, "tmp": tmp, "brand_id": brand_id,
                "date_from": date_from, "date_to": date_to,
            })
            q.put(f"READY:{fid}:{date_from}:{date_to}")
        except Exception as e:
            q.put(f"ERROR:{e}")
            shutil.rmtree(tmp, ignore_errors=True)

    threading.Thread(target=_run, daemon=True).start()

    def _stream():
        while True:
            try:
                msg = q.get(timeout=15)
            except queue.Empty:
                yield ": keepalive\n\n"  # prevent Railway proxy from cutting the connection
                continue
            yield f"data: {msg}\n\n"
            if msg.startswith("READY:") or msg.startswith("ERROR:"):
                break

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/automation/run-chainz-export")
def run_chainz_export_sse(
    _user:    models.User = Depends(get_current_user),
    brand_id: int = Depends(get_brand_id),
):
    """Stream Chainz export automation progress as SSE."""
    with get_db() as db:
        def _s(key):
            r = db.query(models.AppSettings).filter_by(key=key, brand_id=brand_id).first()
            return r.value if r else None
        chainz_email = _s("chainz_email")
        chainz_pass  = _s("chainz_password")
        email_pass   = _s("bosta_email_password")
        gmail_user   = _s("bosta_email")

    if not all([chainz_email, chainz_pass]):
        raise HTTPException(400, "Chainz portal credentials not configured in Settings")
    if not all([gmail_user, email_pass]):
        raise HTTPException(400, "Gmail IMAP credentials (Bosta Integration card) not configured in Settings")

    q = queue.Queue()

    def _run():
        tmp = tempfile.mkdtemp()
        try:
            import chainz_export as ce
            q.put("LOG:Logging in to Chainz portal…")
            ce.trigger_chainz_export(chainz_email, chainz_pass)
            q.put("LOG:Export triggered — waiting for email (up to 5 min)…")
            path = ce.fetch_chainz_email(gmail_user, email_pass)
            q.put("LOG:Attachment downloaded — sorting rows…")
            out, date_from, date_to = ce.sort_chainz_excel(path)
            fid = str(uuid.uuid4())
            _save_export(fid, {
                "path": out, "tmp": tmp, "brand_id": brand_id,
                "date_from": date_from, "date_to": date_to,
                "provider": "chainz",
            })
            q.put(f"READY:{fid}:{date_from}:{date_to}")
        except Exception as e:
            q.put(f"ERROR:{e}")
            shutil.rmtree(tmp, ignore_errors=True)

    threading.Thread(target=_run, daemon=True).start()

    def _stream():
        while True:
            try:
                msg = q.get(timeout=15)
            except queue.Empty:
                yield ": keepalive\n\n"
                continue
            yield f"data: {msg}\n\n"
            if msg.startswith("READY:") or msg.startswith("ERROR:"):
                break

    return StreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class AutomateUploadBody(BaseModel):
    date_from: str
    date_to:   str


@router.post("/automation/upload/{file_id}")
async def automation_upload(
    file_id:  str,
    body:     AutomateUploadBody,
    _user:    models.User = Depends(require_writable),
    brand_id: int = Depends(get_brand_id),
):
    info = _load_export(file_id)
    if not info:
        raise HTTPException(404, "Export session expired or not found")
    if info["brand_id"] != brand_id:
        raise HTTPException(403, "Brand mismatch")
    _delete_export(file_id)
    try:
        with open(info["path"], "rb") as f:
            contents = f.read()
        provider = info.get("provider", "bosta")
        report = await _process_excel(contents, body.date_from, body.date_to, brand_id, provider=provider)
        return JSONResponse(report)
    finally:
        shutil.rmtree(info.get("tmp", ""), ignore_errors=True)


# ── SKU Cost Items ────────────────────────────────────────────────────────────

@router.get("/sku-cost-items")
def get_all_sku_cost_items(
    brand_id: int = Depends(get_brand_id),
    _user: models.User = Depends(get_current_user),
):
    """Return all cost items for this brand, grouped by SKU: { sku: [{name, amount}] }"""
    with get_db() as db:
        rows = db.query(models.SkuCostItem).filter(
            models.SkuCostItem.brand_id == brand_id
        ).all()
    result: dict = {}
    for r in rows:
        result.setdefault(r.sku, []).append({"name": r.name, "amount": r.amount})
    return result


class CostItemIn(BaseModel):
    name:   str
    amount: float

class CostItemsBody(BaseModel):
    items: list[CostItemIn]


@router.put("/sku-cost-items/{sku}")
def upsert_sku_cost_items(
    sku:      str,
    body:     CostItemsBody,
    brand_id: int = Depends(get_brand_id),
    _user:    models.User = Depends(require_writable),
):
    """Full replace: delete existing items for (brand_id, sku), then insert new ones."""
    with get_db() as db:
        db.query(models.SkuCostItem).filter(
            models.SkuCostItem.brand_id == brand_id,
            models.SkuCostItem.sku      == sku,
        ).delete()
        for item in body.items:
            db.add(models.SkuCostItem(
                brand_id=brand_id,
                sku=sku,
                name=item.name.strip(),
                amount=item.amount,
            ))
        db.commit()
    return {"ok": True}
