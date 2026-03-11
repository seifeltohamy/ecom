import re
import json
from collections import defaultdict
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends
from pydantic import BaseModel
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.security import OAuth2PasswordRequestForm
import openpyxl
from sqlalchemy.orm import Session

from app.deps import get_db, get_current_user, require_admin, get_brand_id, oauth2_scheme
from app import models, schemas, auth

app = FastAPI(title="SKU Sales Summary")

PROJECT_DIR = Path(__file__).parent


# ── Auth routes ──────────────────────────────────────────────────────────────

@app.post("/auth/register", tags=["auth"])
def register(
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form("viewer"),
    name: str = Form(""),
    _admin: models.User = Depends(require_admin),
    brand_id: int = Depends(get_brand_id),
):
    if role not in ("admin", "viewer"):
        raise HTTPException(status_code=400, detail="Role must be admin or viewer")
    with get_db() as db:
        if db.query(models.User).filter(models.User.email == email).first():
            raise HTTPException(status_code=400, detail="Email already registered")
        user = models.User(
            email=email,
            password_hash=auth.hash_password(password),
            role=models.UserRole(role),
            name=name.strip() or None,
            brand_id=brand_id if role == "viewer" else None,
        )
        db.add(user)
        db.commit()
    return {"ok": True, "email": email, "role": role}


@app.post("/auth/login", tags=["auth"])
def login(form: OAuth2PasswordRequestForm = Depends()):
    with get_db() as db:
        user = db.query(models.User).filter(models.User.email == form.username).first()
    if not user or not auth.verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    # Admin gets brand_id=None (must pick brand); viewer gets their brand_id
    jwt_brand_id = None if user.role == models.UserRole.admin else user.brand_id
    brand_name = None
    if jwt_brand_id is not None:
        with get_db() as db:
            brand = db.query(models.Brand).filter(models.Brand.id == jwt_brand_id).first()
            brand_name = brand.name if brand else None
    token = auth.create_access_token({
        "sub": user.email,
        "role": user.role.value,
        "brand_id": jwt_brand_id,
        "brand_name": brand_name,
    })
    return {"access_token": token, "token_type": "bearer", "role": user.role.value, "brand_id": jwt_brand_id}


@app.get("/auth/me", tags=["auth"])
def me(token: str = Depends(oauth2_scheme), current_user: models.User = Depends(get_current_user)):
    payload = auth.decode_token(token)
    return {
        "email": current_user.email,
        "role": current_user.role.value,
        "name": current_user.name or "",
        "brand_id": payload.get("brand_id"),
        "brand_name": payload.get("brand_name"),
    }


@app.put("/users/me", tags=["auth"])
def update_my_name(payload: schemas.UserNameUpdate, current_user: models.User = Depends(get_current_user)):
    with get_db() as db:
        user = db.query(models.User).filter(models.User.id == current_user.id).first()
        user.name = payload.name.strip() or None
        db.commit()
    return {"ok": True, "name": payload.name.strip()}


# ── Brand routes ──────────────────────────────────────────────────────────────

@app.get("/brands", tags=["brands"])
def list_brands(_admin: models.User = Depends(require_admin)):
    with get_db() as db:
        brands = db.query(models.Brand).order_by(models.Brand.created_at).all()
        return [{"id": b.id, "name": b.name} for b in brands]


@app.post("/brands", tags=["brands"])
def create_brand(payload: schemas.BrandCreate, _admin: models.User = Depends(require_admin)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Brand name is required.")
    with get_db() as db:
        if db.query(models.Brand).filter(models.Brand.name == name).first():
            raise HTTPException(status_code=400, detail="Brand name already exists.")
        brand = models.Brand(name=name)
        db.add(brand)
        db.commit()
        db.refresh(brand)
        return {"id": brand.id, "name": brand.name}


@app.delete("/brands/{brand_id_param}", tags=["brands"])
def delete_brand(brand_id_param: int, _admin: models.User = Depends(require_admin)):
    with get_db() as db:
        brand = db.query(models.Brand).filter(models.Brand.id == brand_id_param).first()
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found.")
        # Block deletion if any data exists for this brand
        has_data = (
            db.query(models.CashflowMonth).filter(models.CashflowMonth.brand_id == brand_id_param).first()
            or db.query(models.Product).filter(models.Product.brand_id == brand_id_param).first()
            or db.query(models.BostaReport).filter(models.BostaReport.brand_id == brand_id_param).first()
        )
        if has_data:
            raise HTTPException(status_code=400, detail="Cannot delete brand with existing data.")
        db.delete(brand)
        db.commit()
    return {"ok": True}


@app.post("/auth/select-brand", tags=["auth"])
def select_brand(payload: schemas.BrandSelect, current_user: models.User = Depends(require_admin)):
    with get_db() as db:
        brand = db.query(models.Brand).filter(models.Brand.id == payload.brand_id).first()
        if not brand:
            raise HTTPException(status_code=404, detail="Brand not found.")
        brand_name = brand.name
    token = auth.create_access_token({
        "sub": current_user.email,
        "role": "admin",
        "brand_id": payload.brand_id,
        "brand_name": brand_name,
    })
    return {"access_token": token}


@app.post("/auth/clear-brand", tags=["auth"])
def clear_brand(current_user: models.User = Depends(require_admin)):
    """Issue a null-brand admin token so admin returns to the brand picker."""
    token = auth.create_access_token({
        "sub": current_user.email,
        "role": "admin",
        "brand_id": None,
        "brand_name": None,
    })
    return {"access_token": token}


# ── Admin overview ─────────────────────────────────────────────────────────────

@app.get("/admin/overview", tags=["admin"])
def admin_overview(_admin: models.User = Depends(require_admin)):
    """Cross-brand KPI summary for the admin portal. No brand_id required."""
    from datetime import date
    from sqlalchemy import func

    now = date.today()
    current_month_name = now.strftime("%b %Y")  # e.g. "Mar 2026"

    with get_db() as db:
        brands = db.query(models.Brand).order_by(models.Brand.created_at).all()
        result = []
        for brand in brands:
            bid = brand.id

            users_count = db.query(func.count(models.User.id)).filter(
                models.User.brand_id == bid
            ).scalar() or 0

            products_count = db.query(func.count(models.Product.sku)).filter(
                models.Product.brand_id == bid
            ).scalar() or 0

            cashflow_months_count = db.query(func.count(models.CashflowMonth.id)).filter(
                models.CashflowMonth.brand_id == bid
            ).scalar() or 0

            cashflow_entries_total = db.query(func.count(models.CashflowEntry.id)).join(
                models.CashflowMonth, models.CashflowEntry.month_id == models.CashflowMonth.id
            ).filter(models.CashflowMonth.brand_id == bid).scalar() or 0

            # Current month cashflow
            current_month_in = 0.0
            current_month_out = 0.0
            cur_month = db.query(models.CashflowMonth).filter(
                models.CashflowMonth.brand_id == bid,
                models.CashflowMonth.name == current_month_name,
            ).first()
            if cur_month:
                rows = db.query(models.CashflowEntry).filter(
                    models.CashflowEntry.month_id == cur_month.id
                ).all()
                for row in rows:
                    if row.type == "in":
                        current_month_in += row.amount
                    else:
                        current_month_out += row.amount

            bosta_reports_count = db.query(func.count(models.BostaReport.id)).filter(
                models.BostaReport.brand_id == bid
            ).scalar() or 0

            last_report = db.query(models.BostaReport).filter(
                models.BostaReport.brand_id == bid
            ).order_by(models.BostaReport.uploaded_at.desc()).first()
            last_report_date = last_report.uploaded_at.strftime("%b %Y") if last_report else None

            result.append({
                "brand_id": bid,
                "brand_name": brand.name,
                "users_count": users_count,
                "products_count": products_count,
                "cashflow_months_count": cashflow_months_count,
                "cashflow_entries_total": cashflow_entries_total,
                "current_month_in": current_month_in,
                "current_month_out": current_month_out,
                "current_month_net": current_month_in - current_month_out,
                "bosta_reports_count": bosta_reports_count,
                "last_report_date": last_report_date,
            })
        return result


# ── helpers ───────────────────────────────────────────────────────────────────

def get_products_map(db: Session, brand_id: int) -> dict:
    products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
    return {p.sku: p.name for p in products}


def parse_description_text(text: str) -> list[tuple]:
    """Extract (sku, quantity, price) tuples from a Description cell."""
    pattern = r"BostaSKU:(BO-\d+)\s*-\s*quantity:(\d+)\s*-\s*itemPrice:([\d.]+)"
    return re.findall(pattern, text)


def aggregate_excel(workbook, date_from: str = None, date_to: str = None) -> dict:
    from datetime import datetime, date

    ws = workbook.active
    headers = [cell.value for cell in ws[1]]

    try:
        desc_idx = headers.index("Description")
    except ValueError:
        raise HTTPException(status_code=400, detail="No 'Description' column found in Excel.")

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    if "delivered at" in headers_lower:
        deliv_idx = headers_lower.index("delivered at")
    else:
        deliv_idx = next(
            (i for i, h in enumerate(headers_lower) if h == "delivered at"),
            None
        )

    dt_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else None
    dt_to   = datetime.strptime(date_to,   "%Y-%m-%d").date() if date_to   else None

    sku_data = defaultdict(lambda: defaultdict(lambda: {"quantity": 0, "total": 0.0}))
    order_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if deliv_idx is not None and (dt_from or dt_to):
            raw = row[deliv_idx]
            if raw is None:
                continue
            if isinstance(raw, (datetime, date)):
                row_date = raw.date() if isinstance(raw, datetime) else raw
            else:
                raw_str = str(raw).strip()
                parsed = None
                for fmt in ("%m-%d-%Y, %H:%M:%S", "%m-%d-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                    try:
                        parsed = datetime.strptime(raw_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                row_date = parsed
            if dt_from and row_date < dt_from:
                continue
            if dt_to   and row_date > dt_to:
                continue

        cell = row[desc_idx]
        if not cell:
            continue
        matches = parse_description_text(str(cell))
        if matches:
            order_count += 1
            for sku, qty_str, price_str in matches:
                qty = int(qty_str)
                price = float(price_str)
                sku_data[sku][price]["quantity"] += qty
                sku_data[sku][price]["total"] += qty * price

    return sku_data, order_count


def build_report(sku_data: dict, products: dict, order_count: int = 0) -> dict:
    rows = []
    grand_qty = 0
    grand_rev = 0.0

    product_order = {sku: i for i, sku in enumerate(products.keys())}

    def sort_key(sku):
        return (product_order.get(sku, 9999), -sum(v["total"] for v in sku_data[sku].values()))

    for sku in sorted(sku_data.keys(), key=sort_key):
        name = products.get(sku, "Unknown Product")
        price_breakdown = []
        sku_qty = 0
        sku_rev = 0.0

        for price in sorted(sku_data[sku].keys(), reverse=True):
            d = sku_data[sku][price]
            price_breakdown.append({
                "price": price,
                "quantity": d["quantity"],
                "total": round(d["total"], 2),
            })
            sku_qty += d["quantity"]
            sku_rev += d["total"]

        rows.append({
            "sku": sku,
            "name": name,
            "prices": price_breakdown,
            "total_quantity": sku_qty,
            "total_revenue": round(sku_rev, 2),
        })

        grand_qty += sku_qty
        grand_rev += sku_rev

    return {
        "rows": rows,
        "grand_quantity": grand_qty,
        "grand_revenue": round(grand_rev, 2),
        "order_count": order_count,
    }


# routes

@app.post("/debug-upload")
async def debug_upload(file: UploadFile = File(...)):
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    deliv_idx = headers_lower.index("delivered at") if "delivered at" in headers_lower else None

    samples = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= 10:
            break
        raw = row[deliv_idx] if deliv_idx is not None else "column not found"
        samples.append({"raw": str(raw), "type": type(raw).__name__})

    return {"headers": headers, "delivered_at_index": deliv_idx, "samples": samples}


@app.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    date_from: str = Form(None),
    date_to:   str = Form(None),
    brand_id: int = Depends(get_brand_id),
    _user: models.User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls).")

    contents = await file.read()

    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    sku_data, order_count = aggregate_excel(wb, date_from=date_from, date_to=date_to)
    with get_db() as db:
        products = get_products_map(db, brand_id)
    report = build_report(sku_data, products, order_count)

    from datetime import datetime as _dt
    with get_db() as db:
        saved = models.BostaReport(
            uploaded_at    = _dt.utcnow(),
            date_from      = date_from or None,
            date_to        = date_to   or None,
            order_count    = report["order_count"],
            grand_quantity = report["grand_quantity"],
            grand_revenue  = report["grand_revenue"],
            rows_json      = json.dumps(report["rows"]),
            brand_id       = brand_id,
        )
        db.add(saved)
        db.commit()
        db.refresh(saved)
        report["report_id"] = saved.id

    return JSONResponse(report)


@app.get("/products")
def get_products(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
        return {p.sku: p.name for p in products}


@app.post("/products")
def add_product(product: schemas.ProductIn, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    sku = product.sku.strip()
    name = product.name.strip()
    if not sku or not name:
        raise HTTPException(status_code=400, detail="SKU and name are required.")
    with get_db() as db:
        existing = db.query(models.Product).filter(
            models.Product.sku == sku, models.Product.brand_id == brand_id
        ).first()
        if existing:
            existing.name = name
        else:
            db.add(models.Product(sku=sku, name=name, brand_id=brand_id))
        db.commit()
    return {"ok": True, "sku": sku, "name": name}


@app.delete("/products/{sku}")
def delete_product(sku: str, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        product = db.query(models.Product).filter(
            models.Product.sku == sku, models.Product.brand_id == brand_id
        ).first()
        if not product:
            raise HTTPException(status_code=404, detail="SKU not found.")
        db.delete(product)
        db.commit()
    return {"ok": True}


@app.get("/cashflow/months")
def get_cashflow_months(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        months = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.brand_id == brand_id
        ).order_by(models.CashflowMonth.created_at).all()
        return [m.name for m in months]


@app.post("/cashflow/months")
def add_cashflow_month(payload: schemas.CashflowMonthIn, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    month = payload.month.strip()
    if not month:
        raise HTTPException(status_code=400, detail="Month is required.")
    with get_db() as db:
        existing = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not existing:
            db.add(models.CashflowMonth(name=month, brand_id=brand_id))
            db.commit()
        months = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.brand_id == brand_id
        ).order_by(models.CashflowMonth.created_at).all()
        return {"ok": True, "months": [m.name for m in months]}


@app.get("/cashflow/{month}")
def get_cashflow_month(month: str, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            return []
        rows = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id
        ).order_by(models.CashflowEntry.created_at).all()
        return [
            {"id": r.id, "date": r.date, "type": r.type, "amount": r.amount,
             "category": r.category, "notes": r.notes or ""}
            for r in rows
        ]


@app.post("/cashflow/{month}/entries")
def add_cashflow_entry(month: str, entry: schemas.CashflowEntryIn, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            m = models.CashflowMonth(name=month, brand_id=brand_id)
            db.add(m)
            db.commit()
            db.refresh(m)
        row = models.CashflowEntry(
            month_id=m.id, date=entry.date, type=entry.type,
            amount=entry.amount, category=entry.category, notes=entry.notes,
        )
        db.add(row)
        db.commit()
        rows = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id
        ).order_by(models.CashflowEntry.created_at).all()
        return {"ok": True, "rows": [
            {"id": r.id, "date": r.date, "type": r.type, "amount": r.amount,
             "category": r.category, "notes": r.notes or ""}
            for r in rows
        ]}


@app.delete("/cashflow/{month}/entries/{entry_id}")
def delete_cashflow_entry(month: str, entry_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            raise HTTPException(status_code=404, detail="Month not found.")
        row = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id, models.CashflowEntry.id == entry_id
        ).first()
        if not row:
            raise HTTPException(status_code=404, detail="Entry not found.")
        db.add(models.DeletedCashflowEntry(
            id=row.id, month_name=month, date=row.date, type=row.type,
            amount=row.amount, category=row.category, notes=row.notes or "",
            brand_id=brand_id,
        ))
        db.delete(row)
        db.commit()
        rows = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id
        ).order_by(models.CashflowEntry.created_at).all()
        return {"ok": True, "rows": [
            {"id": r.id, "date": r.date, "type": r.type, "amount": r.amount,
             "category": r.category, "notes": r.notes or ""}
            for r in rows
        ]}


@app.put("/cashflow/{month}/entries/{entry_id}")
def update_cashflow_entry(month: str, entry_id: int, entry: schemas.CashflowEntryUpdate, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            raise HTTPException(status_code=404, detail="Month not found.")
        row = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id, models.CashflowEntry.id == entry_id
        ).first()
        if not row:
            raise HTTPException(status_code=404, detail="Entry not found.")
        row.date = entry.date; row.type = entry.type; row.amount = entry.amount
        row.category = entry.category; row.notes = entry.notes
        db.commit()
        rows = db.query(models.CashflowEntry).filter(
            models.CashflowEntry.month_id == m.id
        ).order_by(models.CashflowEntry.created_at).all()
        return {"ok": True, "rows": [
            {"id": r.id, "date": r.date, "type": r.type, "amount": r.amount,
             "category": r.category, "notes": r.notes or ""}
            for r in rows
        ]}


# ── Cashflow categories ───────────────────────────────────────────────────────

@app.get("/categories")
def list_categories(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        cats = db.query(models.CashflowCategory).filter(
            models.CashflowCategory.brand_id == brand_id
        ).order_by(models.CashflowCategory.type, models.CashflowCategory.sort_order, models.CashflowCategory.name).all()
        return [{"id": c.id, "type": c.type, "name": c.name, "sort_order": c.sort_order} for c in cats]


class CategoryIn(BaseModel):
    type: str   # 'in' | 'out'
    name: str

class CategoryReorder(BaseModel):
    ids: list[int]  # ordered list of category ids for one type


@app.post("/categories")
def create_category(payload: CategoryIn, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    if payload.type not in ("in", "out"):
        raise HTTPException(status_code=400, detail="type must be 'in' or 'out'")
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    with get_db() as db:
        existing = db.query(models.CashflowCategory).filter(
            models.CashflowCategory.brand_id == brand_id,
            models.CashflowCategory.type == payload.type,
            models.CashflowCategory.name == name,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail="Category already exists")
        max_order = db.query(models.CashflowCategory).filter(
            models.CashflowCategory.brand_id == brand_id,
            models.CashflowCategory.type == payload.type,
        ).count()
        cat = models.CashflowCategory(brand_id=brand_id, type=payload.type, name=name, sort_order=max_order)
        db.add(cat)
        db.commit()
        return {"id": cat.id, "type": cat.type, "name": cat.name, "sort_order": cat.sort_order}


@app.put("/categories/{cat_id}")
def rename_category(cat_id: int, payload: CategoryIn, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    with get_db() as db:
        cat = db.query(models.CashflowCategory).filter(
            models.CashflowCategory.id == cat_id,
            models.CashflowCategory.brand_id == brand_id,
        ).first()
        if not cat:
            raise HTTPException(status_code=404, detail="Category not found")
        cat.name = name
        db.commit()
        return {"id": cat.id, "type": cat.type, "name": cat.name, "sort_order": cat.sort_order}


@app.delete("/categories/{cat_id}")
def delete_category(cat_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        cat = db.query(models.CashflowCategory).filter(
            models.CashflowCategory.id == cat_id,
            models.CashflowCategory.brand_id == brand_id,
        ).first()
        if not cat:
            raise HTTPException(status_code=404, detail="Category not found")
        db.delete(cat)
        db.commit()
    return {"ok": True}


@app.put("/categories/reorder/{type}")
def reorder_categories(type: str, payload: CategoryReorder, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        for order, cat_id in enumerate(payload.ids):
            db.query(models.CashflowCategory).filter(
                models.CashflowCategory.id == cat_id,
                models.CashflowCategory.brand_id == brand_id,
                models.CashflowCategory.type == type,
            ).update({"sort_order": order})
        db.commit()
    return {"ok": True}


# ── Bosta report history ──────────────────────────────────────────────────────

@app.get("/reports")
def list_reports(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        reports = db.query(models.BostaReport).filter(
            models.BostaReport.brand_id == brand_id
        ).order_by(models.BostaReport.uploaded_at.desc()).all()
        return [
            {
                "id": r.id,
                "uploaded_at": r.uploaded_at.isoformat(),
                "date_from": r.date_from,
                "date_to": r.date_to,
                "order_count": r.order_count,
                "grand_quantity": r.grand_quantity,
                "grand_revenue": r.grand_revenue,
            }
            for r in reports
        ]


@app.get("/reports/{report_id}")
def get_report(report_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        return {
            "report_id": r.id,
            "id": r.id,
            "uploaded_at": r.uploaded_at.isoformat(),
            "date_from": r.date_from,
            "date_to": r.date_to,
            "order_count": r.order_count,
            "grand_quantity": r.grand_quantity,
            "grand_revenue": r.grand_revenue,
            "rows": json.loads(r.rows_json),
        }


@app.get("/reports/{report_id}/pl")
def get_report_pl(report_id: int, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        items = db.query(models.BostaReportPl).filter(
            models.BostaReportPl.report_id == report_id
        ).all()
        return {
            "ads_spent": r.ads_spent,
            "items": [
                {
                    "sku": i.sku, "price": i.price,
                    "cost": i.cost, "extra_cost": i.extra_cost,
                    "cost_formula": i.cost_formula, "extra_cost_formula": i.extra_cost_formula,
                }
                for i in items
            ],
        }


class ReportPlItem(BaseModel):
    sku:                str
    price:              float | None = None
    cost:               float | None = None
    extra_cost:         float | None = None
    cost_formula:       str   | None = None
    extra_cost_formula: str   | None = None

class ReportPlPayload(BaseModel):
    ads_spent: float | None = None
    items:     list[ReportPlItem] = []


@app.put("/reports/{report_id}/pl")
def save_report_pl(report_id: int, payload: ReportPlPayload, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        r = db.query(models.BostaReport).filter(
            models.BostaReport.id == report_id, models.BostaReport.brand_id == brand_id
        ).first()
        if not r:
            raise HTTPException(status_code=404, detail="Report not found.")
        r.ads_spent = payload.ads_spent
        # Upsert per-SKU rows
        for item in payload.items:
            row = db.query(models.BostaReportPl).filter(
                models.BostaReportPl.report_id == report_id,
                models.BostaReportPl.sku == item.sku,
            ).first()
            if row:
                row.price              = item.price
                row.cost               = item.cost
                row.extra_cost         = item.extra_cost
                row.cost_formula       = item.cost_formula
                row.extra_cost_formula = item.extra_cost_formula
            else:
                db.add(models.BostaReportPl(
                    report_id=report_id, sku=item.sku,
                    price=item.price, cost=item.cost, extra_cost=item.extra_cost,
                    cost_formula=item.cost_formula, extra_cost_formula=item.extra_cost_formula,
                ))
        db.commit()
    return {"ok": True}


# ── Dashboard summary ─────────────────────────────────────────────────────────

@app.get("/dashboard/summary")
def dashboard_summary(month: str = None, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    from datetime import datetime as _dt
    now = _dt.utcnow()
    if month:
        current_month_name = month
        current_year = month.split()[-1] if month.split() else str(now.year)
    else:
        current_month_name = now.strftime("%b %Y")
        current_year = str(now.year)

    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == current_month_name,
            models.CashflowMonth.brand_id == brand_id
        ).first()
        this_month_in = this_month_out = 0.0
        if m:
            for e in db.query(models.CashflowEntry).filter(models.CashflowEntry.month_id == m.id).all():
                if e.type == "in":
                    this_month_in  += e.amount
                else:
                    this_month_out += e.amount

        all_months = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.brand_id == brand_id
        ).all()
        ytd_ids = [mm.id for mm in all_months if mm.name.split()[-1] == current_year]
        total_in_ytd = total_out_ytd = 0.0
        if ytd_ids:
            for e in db.query(models.CashflowEntry).filter(models.CashflowEntry.month_id.in_(ytd_ids)).all():
                if e.type == "in":
                    total_in_ytd  += e.amount
                else:
                    total_out_ytd += e.amount

        last_rpt = db.query(models.BostaReport).filter(
            models.BostaReport.brand_id == brand_id
        ).order_by(models.BostaReport.uploaded_at.desc()).first()
        last_report = top_sku = None
        if last_rpt:
            last_report = {
                "uploaded_at": last_rpt.uploaded_at.isoformat(),
                "order_count": last_rpt.order_count,
                "grand_revenue": last_rpt.grand_revenue,
            }
            rows = json.loads(last_rpt.rows_json)
            if rows:
                best = max(rows, key=lambda x: x.get("total_quantity", 0))
                top_sku = {"sku": best["sku"], "name": best["name"], "total_quantity": best["total_quantity"]}

    return {
        "this_month_in":  round(this_month_in,  2),
        "this_month_out": round(this_month_out, 2),
        "this_month_net": round(this_month_in - this_month_out, 2),
        "total_in_ytd":   round(total_in_ytd,   2),
        "total_out_ytd":  round(total_out_ytd,  2),
        "last_report":    last_report,
        "top_sku":        top_sku,
        "current_month":  current_month_name,
    }


# ── User management (admin only) ───────────────────────────────────────────────

@app.get("/users")
def list_users(brand_id: int = Depends(get_brand_id), _admin: models.User = Depends(require_admin)):
    with get_db() as db:
        users = db.query(models.User).filter(
            models.User.brand_id == brand_id
        ).order_by(models.User.created_at).all()
        return [
            {
                "id": u.id,
                "email": u.email,
                "name": u.name or "",
                "role": u.role.value,
                "created_at": u.created_at.isoformat(),
            }
            for u in users
        ]


@app.put("/users/{user_id}", tags=["users"])
def update_user_name(user_id: int, payload: schemas.UserNameUpdate, _admin: models.User = Depends(require_admin)):
    with get_db() as db:
        user = db.query(models.User).filter(models.User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found.")
        user.name = payload.name.strip() or None
        db.commit()
    return {"ok": True, "name": payload.name.strip()}


@app.delete("/users/{user_id}")
def delete_user(user_id: int, admin: models.User = Depends(require_admin)):
    with get_db() as db:
        user = db.query(models.User).filter(models.User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found.")
        if user.id == admin.id:
            raise HTTPException(status_code=400, detail="Cannot delete your own account.")
        db.delete(user)
        db.commit()
    return {"ok": True}


# ── Products Sold ──────────────────────────────────────────────────────────────

def _calc_profit(revenue, cost, qty, extra_cost, expense):
    profit = revenue - (cost or 0) * qty - (extra_cost or 0) - (expense or 0)
    profit_pct = round(profit / revenue * 100, 2) if revenue else 0.0
    return round(profit, 2), profit_pct


@app.get("/products-sold/{month}")
def get_products_sold(month: str, brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    from datetime import datetime as _dt
    try:
        dt = _dt.strptime(month, "%b %Y")
        prefix = dt.strftime("%Y-%m")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format. Use 'Mar 2026'.")

    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            raise HTTPException(status_code=404, detail="Month not found.")

        report = (
            db.query(models.BostaReport)
            .filter(
                models.BostaReport.brand_id == brand_id,
                models.BostaReport.date_from.like(f"{prefix}%"),
            )
            .order_by(models.BostaReport.uploaded_at.desc())
            .first()
        )
        bosta_by_sku = {}
        if report:
            for row in json.loads(report.rows_json):
                bosta_by_sku[row["sku"]] = {
                    "qty": row.get("total_quantity", 0),
                    "revenue": row.get("total_revenue", 0.0),
                }

        products = db.query(models.Product).filter(models.Product.brand_id == brand_id).all()
        manual_rows = {
            r.sku: r
            for r in db.query(models.ProductsSoldManual)
            .filter(models.ProductsSoldManual.month_id == m.id)
            .all()
        }

        result = []
        for p in products:
            b = bosta_by_sku.get(p.sku, {"qty": 0, "revenue": 0.0})
            man = manual_rows.get(p.sku)
            price      = man.price      if man else None
            cost       = man.cost       if man else None
            extra_cost = man.extra_cost if man else None
            expense    = man.expense    if man else None
            profit, profit_pct = _calc_profit(b["revenue"], cost, b["qty"], extra_cost, expense)
            result.append({
                "sku": p.sku, "name": p.name, "price": price, "cost": cost,
                "extra_cost": extra_cost, "qty": b["qty"], "revenue": b["revenue"],
                "expense": expense, "profit": profit, "profit_pct": profit_pct,
            })
        return result


class ProductsSoldUpdate(BaseModel):
    price:      float | None = None
    cost:       float | None = None
    extra_cost: float | None = None
    expense:    float | None = None


@app.put("/products-sold/{month}/{sku}")
def update_products_sold(month: str, sku: str, payload: ProductsSoldUpdate,
                         brand_id: int = Depends(get_brand_id),
                         _user: models.User = Depends(get_current_user)):
    with get_db() as db:
        m = db.query(models.CashflowMonth).filter(
            models.CashflowMonth.name == month, models.CashflowMonth.brand_id == brand_id
        ).first()
        if not m:
            raise HTTPException(status_code=404, detail="Month not found.")

        row = (
            db.query(models.ProductsSoldManual)
            .filter(models.ProductsSoldManual.month_id == m.id,
                    models.ProductsSoldManual.sku == sku)
            .first()
        )
        if row is None:
            row = models.ProductsSoldManual(month_id=m.id, sku=sku)
            db.add(row)

        row.price = payload.price; row.cost = payload.cost
        row.extra_cost = payload.extra_cost; row.expense = payload.expense
        db.commit()
    return {"ok": True}


# ── App Settings ───────────────────────────────────────────────────────────────

class SettingsUpdate(BaseModel):
    bosta_api_key: str | None = None


@app.get("/settings")
def get_settings(brand_id: int = Depends(get_brand_id), _admin: models.User = Depends(require_admin)):
    with get_db() as db:
        rows = db.query(models.AppSettings).filter(models.AppSettings.brand_id == brand_id).all()
        data = {r.key: r.value for r in rows}
    return {"bosta_api_key": data.get("bosta_api_key", "")}


@app.put("/settings")
def update_settings(payload: SettingsUpdate, brand_id: int = Depends(get_brand_id), _admin: models.User = Depends(require_admin)):
    updates = {"bosta_api_key": payload.bosta_api_key or ""}
    with get_db() as db:
        for k, v in updates.items():
            row = db.query(models.AppSettings).filter(
                models.AppSettings.key == k, models.AppSettings.brand_id == brand_id
            ).first()
            if row:
                row.value = v
            else:
                db.add(models.AppSettings(key=k, brand_id=brand_id, value=v))
        db.commit()
    return {"ok": True}


# ── Stock Value ────────────────────────────────────────────────────────────────

@app.get("/stock-value")
def get_stock_value(brand_id: int = Depends(get_brand_id), _user: models.User = Depends(get_current_user)):
    import httpx

    with get_db() as db:
        setting = db.query(models.AppSettings).filter(
            models.AppSettings.key == "bosta_api_key",
            models.AppSettings.brand_id == brand_id,
        ).first()
        api_key = setting.value if setting else ""

    if not api_key:
        raise HTTPException(status_code=400, detail="Bosta API key not configured. Go to Settings to add it.")

    try:
        h = {"Authorization": api_key}
        products = []
        page = 0
        limit = 100
        while True:
            resp = httpx.get(
                "http://app.bosta.co/api/v2/products/fulfillment/list-products",
                headers=h,
                params={"page": page, "limit": limit},
                timeout=15,
                follow_redirects=True,
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"Bosta API returned {resp.status_code}: {resp.text[:500]}")
            data = resp.json().get("data", {})
            batch = data.get("data", [])
            products.extend(batch)
            if len(batch) < limit:
                break
            page += 1

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not reach Bosta API: {str(e)}")

    # Load saved purchase prices for this brand
    with get_db() as db:
        pp_rows = db.query(models.StockPurchasePrice).filter(
            models.StockPurchasePrice.brand_id == brand_id
        ).all()
    purchase_map = {r.sku: r.purchase_price for r in pp_rows}

    rows = []
    for p in products:
        sku            = p.get("product_code") or str(p.get("id", ""))
        name           = p.get("name") or "Unknown"
        consumer_price = p.get("list_price") or 0
        on_hand        = p.get("qty_available") or 0
        reserved       = p.get("virtual_available") or 0
        purchase_price = purchase_map.get(sku, 0)
        rows.append({
            "sku":            sku,
            "name":           name,
            "consumer_price": consumer_price,
            "purchase_price": purchase_price,
            "on_hand":        on_hand,
            "reserved":       reserved,
            "consumer_value": round(on_hand * consumer_price, 2),
            "purchase_value": round(on_hand * purchase_price, 2),
        })

    total_onhand         = sum(r["on_hand"] for r in rows)
    total_consumer_value = round(sum(r["consumer_value"] for r in rows), 2)
    total_purchase_value = round(sum(r["purchase_value"] for r in rows), 2)
    return {
        "rows": rows,
        "total_onhand": total_onhand,
        "total_consumer_value": total_consumer_value,
        "total_purchase_value": total_purchase_value,
    }


class StockPurchasePriceIn(BaseModel):
    sku:   str
    price: float


@app.put("/stock-value/purchase-price")
def upsert_purchase_price(
    body:     StockPurchasePriceIn,
    brand_id: int = Depends(get_brand_id),
    _user:    models.User = Depends(get_current_user),
):
    with get_db() as db:
        row = db.query(models.StockPurchasePrice).filter(
            models.StockPurchasePrice.brand_id == brand_id,
            models.StockPurchasePrice.sku      == body.sku,
        ).first()
        if row:
            row.purchase_price = body.price
        else:
            db.add(models.StockPurchasePrice(
                brand_id=brand_id, sku=body.sku, purchase_price=body.price
            ))
        db.commit()
    return {"ok": True}


# ── SPA / static file serving ────────────────────────────────────────────────

from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

DIST = PROJECT_DIR / "frontend" / "dist"
if DIST.exists():
    app.mount("/assets", StaticFiles(directory=str(DIST / "assets")), name="assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def spa_fallback(full_path: str = ""):  # noqa: ARG001
        return FileResponse(str(DIST / "index.html"))
else:
    @app.get("/", response_class=HTMLResponse)
    def serve_dev():
        return HTMLResponse((PROJECT_DIR / "index.html").read_text(encoding="utf-8"))
